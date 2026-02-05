"""
Check the Scenario Planner sheet for errors
"""

from openpyxl import load_workbook

wb = load_workbook('C:/Users/xavie/Documents/mushroom_farm/mushroom_calculator.xlsx')

ws = wb['Scenario Planner']

print("=== SCENARIO PLANNER SHEET ===")
for row in range(1, 30):
    row_data = []
    for col in range(1, 7):
        cell = ws.cell(row=row, column=col)
        val = cell.value
        if val is not None:
            row_data.append(f"[{col}]{val}")
    if row_data:
        print(f"Row {row}: {' | '.join(row_data)}")

print("\n=== CHECKING REFERENCED CELLS IN OYSTER ===")
ws_oyster = wb['Oyster']
print(f"B32 (was yield per bag): {ws_oyster['B32'].value}")
print(f"B33 (was cycle length): {ws_oyster['B33'].value}")
print(f"B20 (was sale price): {ws_oyster['B20'].value}")
print(f"B17 (was cost per bag): {ws_oyster['B17'].value}")

print("\n=== ACTUAL CELL LOCATIONS IN OYSTER ===")
print(f"B26 (yield per bag): {ws_oyster['B26'].value}")
print(f"B27 (cycle length): {ws_oyster['B27'].value}")
print(f"B22 (sale price): {ws_oyster['B22'].value}")
print(f"B19 (total cost per bag): {ws_oyster['B19'].value}")

print("\n=== CHECKING REFERENCED CELLS IN LIONS MANE ===")
ws_lm = wb['Lions Mane']
print(f"B32: {ws_lm['B32'].value}")
print(f"B33: {ws_lm['B33'].value}")
print(f"B20: {ws_lm['B20'].value}")
print(f"B17: {ws_lm['B17'].value}")

print("\n=== ACTUAL CELL LOCATIONS IN LIONS MANE ===")
print(f"B26 (yield per bag): {ws_lm['B26'].value}")
print(f"B27 (cycle length): {ws_lm['B27'].value}")
print(f"B22 (sale price): {ws_lm['B22'].value}")
print(f"B19 (total cost per bag): {ws_lm['B19'].value}")

wb.close()
