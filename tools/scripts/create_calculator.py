"""
Mushroom Production & Financial Calculator - Excel Generator
Creates an xlsx file with formulas for production planning
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def create_calculator():
    wb = Workbook()

    # === OYSTER SHEET ===
    ws_oyster = wb.active
    ws_oyster.title = "Oyster"
    setup_variety_sheet(ws_oyster, {
        'name': 'OYSTER MUSHROOMS',
        'target_kg': 20,
        'bag_weight': 5,
        'dry_pct': 25,
        'be_pct': 100,
        'incubation_weeks': 2.5,
        'fruiting_weeks': 2.5,
        'substrate_cost': 2.50,
        'spawn_cost': 1.50,
        'bag_cost': 0.80,
        'labor_cost': 1.00,
        'utilities_cost': 0.50,
        'sale_price': 25.00
    })

    # === LION'S MANE SHEET ===
    ws_lions = wb.create_sheet("Lions Mane")
    setup_variety_sheet(ws_lions, {
        'name': "LION'S MANE",
        'target_kg': 10,
        'bag_weight': 5,
        'dry_pct': 25,
        'be_pct': 80,
        'incubation_weeks': 3.5,
        'fruiting_weeks': 3.5,
        'substrate_cost': 3.00,
        'spawn_cost': 2.00,
        'bag_cost': 0.80,
        'labor_cost': 1.50,
        'utilities_cost': 1.00,
        'sale_price': 35.00
    })

    # === SUMMARY SHEET ===
    ws_summary = wb.create_sheet("Summary")
    setup_summary_sheet(ws_summary)

    # === SCENARIO PLANNER ===
    ws_scenario = wb.create_sheet("Scenario Planner")
    setup_scenario_sheet(ws_scenario)

    wb.save('C:/Users/xavie/Documents/mushroom_farm/mushroom_calculator.xlsx')
    print("Created: mushroom_calculator.xlsx")

# Styles
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
calc_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
money_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
label_font = Font(bold=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def setup_variety_sheet(ws, config):
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 40

    row = 1

    # Title
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = f"{config['name']} - PRODUCTION CALCULATOR"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
    row += 2

    # --- PRODUCTION INPUTS ---
    ws[f'A{row}'] = "PRODUCTION INPUTS"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    # Headers
    for col, header in enumerate(['Parameter', 'Value', 'Unit', 'Notes'], 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = label_font
        cell.border = thin_border
    row += 1

    # Input rows with yellow highlighting
    inputs = [
        ('Target weekly output', config['target_kg'], 'kg', 'Your weekly sales target'),
        ('Substrate bag weight', config['bag_weight'], 'kg', 'Hydrated weight of prepared bag'),
        ('Dry substrate %', config['dry_pct'], '%', 'Typical 25% for straw/sawdust'),
        ('Biological efficiency', config['be_pct'], '%', 'Oyster 100-150%, LM 75-100%'),
        ('Incubation period', config['incubation_weeks'], 'weeks', 'Time in incubation'),
        ('Fruiting period', config['fruiting_weeks'], 'weeks', 'Time for 2 flushes'),
    ]

    input_start_row = row
    for label, value, unit, notes in inputs:
        ws.cell(row=row, column=1, value=label).border = thin_border
        cell = ws.cell(row=row, column=2, value=value)
        cell.fill = input_fill
        cell.border = thin_border
        ws.cell(row=row, column=3, value=unit).border = thin_border
        ws.cell(row=row, column=4, value=notes).border = thin_border
        row += 1

    row += 1

    # --- COST INPUTS ---
    ws[f'A{row}'] = "COST INPUTS (AUD)"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    for col, header in enumerate(['Cost Item', 'Amount', 'Unit', 'Notes'], 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = label_font
        cell.border = thin_border
    row += 1

    cost_start_row = row
    costs = [
        ('Substrate per bag', config['substrate_cost'], '$/bag', 'Straw, sawdust, supplements'),
        ('Spawn per bag', config['spawn_cost'], '$/bag', 'Grain spawn allocation'),
        ('Bag & filter', config['bag_cost'], '$/bag', 'Autoclavable grow bag'),
        ('Labor per bag', config['labor_cost'], '$/bag', 'Your time value'),
        ('Utilities per bag', config['utilities_cost'], '$/bag', 'Power, water, cooling'),
    ]

    for label, value, unit, notes in costs:
        ws.cell(row=row, column=1, value=label).border = thin_border
        cell = ws.cell(row=row, column=2, value=value)
        cell.fill = input_fill
        cell.border = thin_border
        cell.number_format = '$#,##0.00'
        ws.cell(row=row, column=3, value=unit).border = thin_border
        ws.cell(row=row, column=4, value=notes).border = thin_border
        row += 1

    # Total cost formula
    ws.cell(row=row, column=1, value="TOTAL COST PER BAG").font = label_font
    ws.cell(row=row, column=1).border = thin_border
    cost_end_row = row - 1
    cell = ws.cell(row=row, column=2, value=f"=SUM(B{cost_start_row}:B{cost_end_row})")
    cell.fill = calc_fill
    cell.border = thin_border
    cell.number_format = '$#,##0.00'
    cell.font = label_font
    ws.cell(row=row, column=3, value='$/bag').border = thin_border
    ws.cell(row=row, column=4, value='=SUM of all costs above').border = thin_border
    total_cost_row = row
    row += 2

    # --- REVENUE INPUT ---
    ws[f'A{row}'] = "REVENUE"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    ws.cell(row=row, column=1, value="Sale price per kg").border = thin_border
    cell = ws.cell(row=row, column=2, value=config['sale_price'])
    cell.fill = input_fill
    cell.border = thin_border
    cell.number_format = '$#,##0.00'
    ws.cell(row=row, column=3, value='$/kg').border = thin_border
    ws.cell(row=row, column=4, value='Farm gate price').border = thin_border
    sale_price_row = row
    row += 2

    # --- CALCULATED PRODUCTION ---
    ws[f'A{row}'] = "CALCULATED PRODUCTION"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    # Dry substrate per bag
    ws.cell(row=row, column=1, value="Dry substrate per bag").border = thin_border
    cell = ws.cell(row=row, column=2, value=f"=B{input_start_row+1}*B{input_start_row+2}/100")
    cell.fill = calc_fill
    cell.border = thin_border
    cell.number_format = '0.00'
    ws.cell(row=row, column=3, value='kg').border = thin_border
    dry_sub_row = row
    row += 1

    # Yield per bag
    ws.cell(row=row, column=1, value="Yield per bag").border = thin_border
    cell = ws.cell(row=row, column=2, value=f"=B{dry_sub_row}*B{input_start_row+3}/100")
    cell.fill = calc_fill
    cell.border = thin_border
    cell.number_format = '0.00'
    ws.cell(row=row, column=3, value='kg').border = thin_border
    ws.cell(row=row, column=4, value='Fresh mushrooms from 2 flushes').border = thin_border
    yield_row = row
    row += 1

    # Cycle length
    ws.cell(row=row, column=1, value="Total cycle length").border = thin_border
    cell = ws.cell(row=row, column=2, value=f"=B{input_start_row+4}+B{input_start_row+5}")
    cell.fill = calc_fill
    cell.border = thin_border
    cell.number_format = '0.0'
    ws.cell(row=row, column=3, value='weeks').border = thin_border
    cycle_row = row
    row += 1

    # Bags per week
    ws.cell(row=row, column=1, value="Bags to prep per week").border = thin_border
    cell = ws.cell(row=row, column=2, value=f"=CEILING(B{input_start_row}/B{yield_row},1)")
    cell.fill = calc_fill
    cell.border = thin_border
    cell.number_format = '0'
    ws.cell(row=row, column=3, value='bags').border = thin_border
    ws.cell(row=row, column=4, value='Bags to pasteurize weekly').border = thin_border
    bags_week_row = row
    row += 1

    # Total in rotation
    ws.cell(row=row, column=1, value="Total bags in rotation").border = thin_border
    cell = ws.cell(row=row, column=2, value=f"=B{bags_week_row}*B{cycle_row}")
    cell.fill = calc_fill
    cell.border = thin_border
    cell.number_format = '0'
    ws.cell(row=row, column=3, value='bags').border = thin_border
    ws.cell(row=row, column=4, value='All bags in system at once').border = thin_border
    total_bags_row = row
    row += 2

    # --- FINANCIALS ---
    ws[f'A{row}'] = "FINANCIAL RESULTS"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    for col, header in enumerate(['Metric', 'Weekly', 'Monthly', 'Yearly'], 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = label_font
        cell.border = thin_border
    row += 1

    # Gross Revenue
    ws.cell(row=row, column=1, value="Gross Revenue").border = thin_border
    cell = ws.cell(row=row, column=2, value=f"=B{input_start_row}*B{sale_price_row}")
    cell.fill = money_fill
    cell.border = thin_border
    cell.number_format = '$#,##0.00'
    cell = ws.cell(row=row, column=3, value=f"=B{row}*4.33")
    cell.fill = money_fill
    cell.border = thin_border
    cell.number_format = '$#,##0.00'
    cell = ws.cell(row=row, column=4, value=f"=B{row}*52")
    cell.fill = money_fill
    cell.border = thin_border
    cell.number_format = '$#,##0.00'
    revenue_row = row
    row += 1

    # Total Costs
    ws.cell(row=row, column=1, value="Total Costs").border = thin_border
    cell = ws.cell(row=row, column=2, value=f"=B{bags_week_row}*B{total_cost_row}")
    cell.fill = money_fill
    cell.border = thin_border
    cell.number_format = '$#,##0.00'
    cell = ws.cell(row=row, column=3, value=f"=B{row}*4.33")
    cell.fill = money_fill
    cell.border = thin_border
    cell.number_format = '$#,##0.00'
    cell = ws.cell(row=row, column=4, value=f"=B{row}*52")
    cell.fill = money_fill
    cell.border = thin_border
    cell.number_format = '$#,##0.00'
    costs_row = row
    row += 1

    # Gross Profit
    ws.cell(row=row, column=1, value="GROSS PROFIT").font = label_font
    ws.cell(row=row, column=1).border = thin_border
    cell = ws.cell(row=row, column=2, value=f"=B{revenue_row}-B{costs_row}")
    cell.fill = calc_fill
    cell.border = thin_border
    cell.number_format = '$#,##0.00'
    cell.font = label_font
    cell = ws.cell(row=row, column=3, value=f"=C{revenue_row}-C{costs_row}")
    cell.fill = calc_fill
    cell.border = thin_border
    cell.number_format = '$#,##0.00'
    cell.font = label_font
    cell = ws.cell(row=row, column=4, value=f"=D{revenue_row}-D{costs_row}")
    cell.fill = calc_fill
    cell.border = thin_border
    cell.number_format = '$#,##0.00'
    cell.font = label_font
    profit_row = row
    row += 1

    # Profit Margin
    ws.cell(row=row, column=1, value="Profit Margin").border = thin_border
    cell = ws.cell(row=row, column=2, value=f"=B{profit_row}/B{revenue_row}")
    cell.fill = calc_fill
    cell.border = thin_border
    cell.number_format = '0.0%'
    row += 2

    # --- KEY METRICS ---
    ws[f'A{row}'] = "KEY METRICS"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    ws.cell(row=row, column=1, value="Cost per kg produced").border = thin_border
    cell = ws.cell(row=row, column=2, value=f"=B{total_cost_row}/B{yield_row}")
    cell.fill = calc_fill
    cell.border = thin_border
    cell.number_format = '$#,##0.00'
    ws.cell(row=row, column=3, value='$/kg').border = thin_border
    ws.cell(row=row, column=4, value='Your breakeven price').border = thin_border
    row += 1

    ws.cell(row=row, column=1, value="Profit per kg").border = thin_border
    cell = ws.cell(row=row, column=2, value=f"=B{sale_price_row}-(B{total_cost_row}/B{yield_row})")
    cell.fill = calc_fill
    cell.border = thin_border
    cell.number_format = '$#,##0.00'
    ws.cell(row=row, column=3, value='$/kg').border = thin_border
    row += 1

    ws.cell(row=row, column=1, value="Profit per bag").border = thin_border
    cell = ws.cell(row=row, column=2, value=f"=(B{sale_price_row}*B{yield_row})-B{total_cost_row}")
    cell.fill = calc_fill
    cell.border = thin_border
    cell.number_format = '$#,##0.00'
    ws.cell(row=row, column=3, value='$/bag').border = thin_border


def setup_summary_sheet(ws):
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    row = 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "COMBINED OPERATION SUMMARY"
    ws[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws[f'A{row}'].fill = header_fill
    row += 2

    headers = ['Metric', 'Oyster', "Lion's Mane", 'TOTAL']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = label_font
        cell.border = thin_border
    row += 1

    # References to other sheets
    metrics = [
        ('Target kg/week', "=Oyster!B4", "='Lions Mane'!B4", "=B{r}+C{r}"),
        ('Yield per bag (kg)', "=Oyster!B32", "='Lions Mane'!B32", "-"),
        ('Bags per week', "=Oyster!B35", "='Lions Mane'!B35", "=B{r}+C{r}"),
        ('Bags in rotation', "=Oyster!B36", "='Lions Mane'!B36", "=B{r}+C{r}"),
        ('Cost per bag', "=Oyster!B17", "='Lions Mane'!B17", "-"),
        ('Sale price/kg', "=Oyster!B20", "='Lions Mane'!B20", "-"),
    ]

    for label, oyster_ref, lions_ref, total_formula in metrics:
        ws.cell(row=row, column=1, value=label).border = thin_border
        cell = ws.cell(row=row, column=2, value=oyster_ref)
        cell.border = thin_border
        if '$' in label or 'Cost' in label or 'price' in label:
            cell.number_format = '$#,##0.00'
        else:
            cell.number_format = '0.0'
        cell = ws.cell(row=row, column=3, value=lions_ref)
        cell.border = thin_border
        if '$' in label or 'Cost' in label or 'price' in label:
            cell.number_format = '$#,##0.00'
        else:
            cell.number_format = '0.0'
        if total_formula != "-":
            cell = ws.cell(row=row, column=4, value=total_formula.format(r=row))
            cell.fill = calc_fill
            cell.number_format = '0.0'
        cell.border = thin_border
        row += 1

    row += 1
    ws[f'A{row}'] = "FINANCIAL SUMMARY"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    for col, header in enumerate(['Metric', 'Oyster', "Lion's Mane", 'TOTAL'], 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = label_font
        cell.border = thin_border
    row += 1

    financials = [
        ('Weekly Revenue', "=Oyster!B40", "='Lions Mane'!B40"),
        ('Weekly Costs', "=Oyster!B41", "='Lions Mane'!B41"),
        ('Weekly Profit', "=Oyster!B42", "='Lions Mane'!B42"),
        ('Monthly Profit', "=Oyster!C42", "='Lions Mane'!C42"),
        ('Yearly Profit', "=Oyster!D42", "='Lions Mane'!D42"),
    ]

    for label, oyster_ref, lions_ref in financials:
        ws.cell(row=row, column=1, value=label).border = thin_border
        cell = ws.cell(row=row, column=2, value=oyster_ref)
        cell.border = thin_border
        cell.number_format = '$#,##0.00'
        cell.fill = money_fill
        cell = ws.cell(row=row, column=3, value=lions_ref)
        cell.border = thin_border
        cell.number_format = '$#,##0.00'
        cell.fill = money_fill
        cell = ws.cell(row=row, column=4, value=f"=B{row}+C{row}")
        cell.border = thin_border
        cell.number_format = '$#,##0.00'
        cell.fill = calc_fill
        if 'Profit' in label:
            cell.font = label_font
        row += 1

    row += 2
    ws[f'A{row}'] = "SPACE CHECK - FRUITING ROOM"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    ws.cell(row=row, column=1, value="Room size (sqm)").border = thin_border
    cell = ws.cell(row=row, column=2, value=9)
    cell.fill = input_fill
    cell.border = thin_border
    room_size_row = row
    row += 1

    ws.cell(row=row, column=1, value="Bags per sqm (4-tier)").border = thin_border
    cell = ws.cell(row=row, column=2, value=12)
    cell.fill = input_fill
    cell.border = thin_border
    bags_sqm_row = row
    row += 1

    ws.cell(row=row, column=1, value="Max room capacity").border = thin_border
    cell = ws.cell(row=row, column=2, value=f"=B{room_size_row}*B{bags_sqm_row}")
    cell.fill = calc_fill
    cell.border = thin_border
    cell.number_format = '0'
    ws.cell(row=row, column=3, value="bags").border = thin_border
    capacity_row = row
    row += 1

    ws.cell(row=row, column=1, value="Bags in fruiting (~50%)").border = thin_border
    cell = ws.cell(row=row, column=2, value=f"=D8*0.5")
    cell.fill = calc_fill
    cell.border = thin_border
    cell.number_format = '0'
    ws.cell(row=row, column=3, value="bags").border = thin_border
    in_fruiting_row = row
    row += 1

    ws.cell(row=row, column=1, value="Space utilization").border = thin_border
    cell = ws.cell(row=row, column=2, value=f"=B{in_fruiting_row}/B{capacity_row}")
    cell.fill = calc_fill
    cell.border = thin_border
    cell.number_format = '0.0%'
    row += 1

    ws.cell(row=row, column=1, value="Status").border = thin_border
    cell = ws.cell(row=row, column=2, value=f'=IF(B{in_fruiting_row}<=B{capacity_row},"OK - FITS","WARNING - OVER CAPACITY")')
    cell.fill = calc_fill
    cell.border = thin_border
    cell.font = label_font


def setup_scenario_sheet(ws):
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12

    row = 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "SCENARIO PLANNER - What if I produce X kg/week?"
    ws[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws[f'A{row}'].fill = header_fill
    row += 2

    ws[f'A{row}'] = "Oyster - vary target output"
    ws[f'A{row}'].font = label_font
    row += 1

    headers = ['Target kg/wk', 'Bags/wk', 'In Rotation', 'Weekly Profit', 'Monthly Profit', 'Yearly Profit']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = label_font
        cell.border = thin_border
    row += 1

    targets = [5, 10, 15, 20, 25, 30, 40, 50]
    for target in targets:
        ws.cell(row=row, column=1, value=target).border = thin_border
        # Bags/week = target / yield per bag (using Oyster sheet values)
        cell = ws.cell(row=row, column=2, value=f"=CEILING(A{row}/Oyster!B32,1)")
        cell.border = thin_border
        cell.number_format = '0'
        # In rotation
        cell = ws.cell(row=row, column=3, value=f"=B{row}*Oyster!B33")
        cell.border = thin_border
        cell.number_format = '0'
        # Weekly profit
        cell = ws.cell(row=row, column=4, value=f"=(A{row}*Oyster!B20)-(B{row}*Oyster!B17)")
        cell.border = thin_border
        cell.number_format = '$#,##0'
        cell.fill = money_fill
        # Monthly
        cell = ws.cell(row=row, column=5, value=f"=D{row}*4.33")
        cell.border = thin_border
        cell.number_format = '$#,##0'
        cell.fill = money_fill
        # Yearly
        cell = ws.cell(row=row, column=6, value=f"=D{row}*52")
        cell.border = thin_border
        cell.number_format = '$#,##0'
        cell.fill = money_fill
        row += 1

    row += 2
    ws[f'A{row}'] = "Lion's Mane - vary target output"
    ws[f'A{row}'].font = label_font
    row += 1

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = label_font
        cell.border = thin_border
    row += 1

    targets_lm = [3, 5, 8, 10, 12, 15, 20]
    for target in targets_lm:
        ws.cell(row=row, column=1, value=target).border = thin_border
        cell = ws.cell(row=row, column=2, value=f"=CEILING(A{row}/'Lions Mane'!B32,1)")
        cell.border = thin_border
        cell.number_format = '0'
        cell = ws.cell(row=row, column=3, value=f"=B{row}*'Lions Mane'!B33")
        cell.border = thin_border
        cell.number_format = '0'
        cell = ws.cell(row=row, column=4, value=f"=(A{row}*'Lions Mane'!B20)-(B{row}*'Lions Mane'!B17)")
        cell.border = thin_border
        cell.number_format = '$#,##0'
        cell.fill = money_fill
        cell = ws.cell(row=row, column=5, value=f"=D{row}*4.33")
        cell.border = thin_border
        cell.number_format = '$#,##0'
        cell.fill = money_fill
        cell = ws.cell(row=row, column=6, value=f"=D{row}*52")
        cell.border = thin_border
        cell.number_format = '$#,##0'
        cell.fill = money_fill
        row += 1


if __name__ == "__main__":
    create_calculator()
