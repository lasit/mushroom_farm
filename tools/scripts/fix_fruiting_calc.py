"""
Fix the bags in fruiting calculation to be dynamic based on actual cycle times
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the workbook
wb = load_workbook('C:/Users/xavie/Documents/mushroom_farm/mushroom_calculator.xlsx')

ws = wb['Summary']

# Fix Row 24: Bags in fruiting
# Old formula: =D7*0.5 (hardcoded 50%)
# New formula: Calculate each variety's bags in fruiting based on their actual fruiting/cycle ratio
#
# Oyster: B29 = bags in rotation, B10 = fruiting weeks, B27 = total cycle
# Lions Mane: B29 = bags in rotation, B10 = fruiting weeks, B27 = total cycle
#
# Bags in fruiting = (Oyster bags × Oyster fruiting%) + (LM bags × LM fruiting%)

new_formula = "=(Oyster!B29 * Oyster!B10 / Oyster!B27) + ('Lions Mane'!B29 * 'Lions Mane'!B10 / 'Lions Mane'!B27)"

# Update the cell
ws['B24'] = new_formula
ws['B24'].number_format = '0'

# Update the label to remove the hardcoded hint
ws['A24'] = "Bags in fruiting"

# Also add a calc fill to show it's calculated
calc_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
ws['B24'].fill = calc_fill

print("Fixed Row 24:")
print(f"  Old: =D7*0.5")
print(f"  New: {new_formula}")
print()
print("This now calculates:")
print("  Oyster bags in fruiting = Oyster rotation × (Oyster fruiting weeks ÷ Oyster cycle)")
print("  + Lion's Mane bags in fruiting = LM rotation × (LM fruiting weeks ÷ LM cycle)")

# Save the workbook
wb.save('C:/Users/xavie/Documents/mushroom_farm/mushroom_calculator.xlsx')
print()
print("Saved: mushroom_calculator.xlsx")

wb.close()
