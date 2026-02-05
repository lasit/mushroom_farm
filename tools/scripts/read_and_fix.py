"""
Read the current spreadsheet structure and fix the bags in fruiting calculation
"""

from openpyxl import load_workbook

# Load the user's updated workbook
wb = load_workbook('C:/Users/xavie/Documents/mushroom_farm/mushroom_calculator.xlsx')

print("=== SHEETS ===")
print(wb.sheetnames)

# Check each variety sheet to find the relevant cell references
for sheet_name in ['Oyster', 'Lions Mane']:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n=== {sheet_name} SHEET ===")

        # Print all rows with values to understand structure
        for row in range(1, 60):
            row_data = []
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                val = cell.value
                if val is not None:
                    row_data.append(f"[{col}]{val}")
            if row_data:
                print(f"Row {row}: {' | '.join(row_data)}")

# Check summary sheet
if 'Summary' in wb.sheetnames:
    ws = wb['Summary']
    print(f"\n=== SUMMARY SHEET ===")
    for row in range(1, 35):
        row_data = []
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if val is not None:
                row_data.append(f"[{col}]{val}")
        if row_data:
            print(f"Row {row}: {' | '.join(row_data)}")

wb.close()
