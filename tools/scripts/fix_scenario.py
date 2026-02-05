"""
Fix the Scenario Planner sheet - correct all cell references
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side

wb = load_workbook('C:/Users/xavie/Documents/mushroom_farm/mushroom_calculator.xlsx')
ws = wb['Scenario Planner']

money_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Correct references:
# Yield per bag: B26 (not B32)
# Cycle length: B27 (not B33)
# Sale price: B22 (not B20)
# Total cost per bag: B19 (not B17)

print("Fixing Oyster scenario rows (5-12)...")
for row in range(5, 13):
    # Column B: Bags/wk = CEILING(target / yield per bag, 1)
    ws.cell(row=row, column=2, value=f"=CEILING(A{row}/Oyster!B26,1)")
    ws.cell(row=row, column=2).border = thin_border
    ws.cell(row=row, column=2).number_format = '0'

    # Column C: In Rotation = bags/wk * cycle length
    ws.cell(row=row, column=3, value=f"=B{row}*Oyster!B27")
    ws.cell(row=row, column=3).border = thin_border
    ws.cell(row=row, column=3).number_format = '0'

    # Column D: Weekly Profit = (target * sale price) - (bags/wk * cost per bag)
    ws.cell(row=row, column=4, value=f"=(A{row}*Oyster!B22)-(B{row}*Oyster!B19)")
    ws.cell(row=row, column=4).border = thin_border
    ws.cell(row=row, column=4).number_format = '$#,##0'
    ws.cell(row=row, column=4).fill = money_fill

    # Column E: Monthly = weekly * 4.33
    ws.cell(row=row, column=5, value=f"=D{row}*4.33")
    ws.cell(row=row, column=5).border = thin_border
    ws.cell(row=row, column=5).number_format = '$#,##0'
    ws.cell(row=row, column=5).fill = money_fill

    # Column F: Yearly = weekly * 52
    ws.cell(row=row, column=6, value=f"=D{row}*52")
    ws.cell(row=row, column=6).border = thin_border
    ws.cell(row=row, column=6).number_format = '$#,##0'
    ws.cell(row=row, column=6).fill = money_fill

print("Fixing Lion's Mane scenario rows (17-23)...")
for row in range(17, 24):
    # Column B: Bags/wk
    ws.cell(row=row, column=2, value=f"=CEILING(A{row}/'Lions Mane'!B26,1)")
    ws.cell(row=row, column=2).border = thin_border
    ws.cell(row=row, column=2).number_format = '0'

    # Column C: In Rotation
    ws.cell(row=row, column=3, value=f"=B{row}*'Lions Mane'!B27")
    ws.cell(row=row, column=3).border = thin_border
    ws.cell(row=row, column=3).number_format = '0'

    # Column D: Weekly Profit
    ws.cell(row=row, column=4, value=f"=(A{row}*'Lions Mane'!B22)-(B{row}*'Lions Mane'!B19)")
    ws.cell(row=row, column=4).border = thin_border
    ws.cell(row=row, column=4).number_format = '$#,##0'
    ws.cell(row=row, column=4).fill = money_fill

    # Column E: Monthly
    ws.cell(row=row, column=5, value=f"=D{row}*4.33")
    ws.cell(row=row, column=5).border = thin_border
    ws.cell(row=row, column=5).number_format = '$#,##0'
    ws.cell(row=row, column=5).fill = money_fill

    # Column F: Yearly
    ws.cell(row=row, column=6, value=f"=D{row}*52")
    ws.cell(row=row, column=6).border = thin_border
    ws.cell(row=row, column=6).number_format = '$#,##0'
    ws.cell(row=row, column=6).fill = money_fill

wb.save('C:/Users/xavie/Documents/mushroom_farm/mushroom_calculator.xlsx')
print("\nFixed all references:")
print("  B32 -> B26 (yield per bag)")
print("  B33 -> B27 (cycle length)")
print("  B20 -> B22 (sale price)")
print("  B17 -> B19 (total cost per bag)")
print("\nSaved: mushroom_calculator.xlsx")

wb.close()
